# !/usr/bin python3                                 
# encoding: utf-8 -*-                            
# @author: 沙陌 微信：Matongxue_2
# @Time: 2022-02-19 15:06
# @Copyright：北京码同学

import openpyxl

# 你要读哪一个文件，以及要读excel里的哪个sheet工作表，作为参数传给该方法
from requests_client import send


def read_excel(file_path, sheet_name):
    sheet_data = openpyxl.load_workbook(file_path)[sheet_name]  # 获取excel中某个sheet工作表的内容
    # 遍历excel数据，逐行逐列读取
    # 需要知道excel有多少行，有多少列
    lines_count = sheet_data.max_row  # 获取最大行数
    cols_count = sheet_data.max_column  # 获取最大列数
    # 读取的数据需要提供给调用方使用，所以我们要想一个数据的存储变量来存储
    data = []
    for r in range(2, lines_count + 1):
        # 把每一行数据都存在列表中
        lines = []
        for c in range(1, cols_count + 1):
            cell_value = sheet_data.cell(r, c).value
            lines.append(cell_value)
        # 当前行所有数据读取完成后，会存入lines，再把lines存入到data里
        data.append(lines)
    return data


if __name__ == '__main__':
    data = read_excel('shop.xlsx', 'Sheet2')
    # ['登录接口', 'http://121.42.15.146:9090/mtx/index.php?s=/index/user/login.html', 'post',
    # '{\n"X-Requested-With":"XMLHttpRequest"\n}', '{\n"data":{\n"accounts":"shamotest1",\n"pwd":"123456"\n}\n}',
    # 200, 0, '登录成功']
    for reqs in data:
        casename = reqs[0]
        url = reqs[1]
        method = reqs[2]
        headers = eval(reqs[3])
        params = eval(reqs[4])  # 把json格式的字符串转换成python里字典对象
        expect_status = reqs[5]
        # exTypeError: send() argument after ** must be a mapping, not strpect_status = reqs[5] #期望的响应状态码
        expect_busi_status = reqs[6] #期望的响应业务码
        expect_msg = reqs[7] # 期望的提示语
        resp = send(url=url, method=method, headers=headers, **params)
        print(resp.text)
        actual_status= resp.status_code #实际的响应状态码
        actual_busi_status = resp.json()['code'] #从响应中提取code值
        actual_msg = resp.json()['msg']
        if expect_status == actual_status:
            print('响应状态码正确')
        else:
            print('响应状态码不正确')
